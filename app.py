"""
app.py

A minimal-threshold Flask application for Association Analysis (Apriori):
 - min_support, min_confidence, min_combination_count
 - Summation of combination counts to determine node size
 - Excel and PNG download
 - Server-side session (Flask-Session)

Licensed under GNU GPL (v3 or later): https://www.gnu.org/licenses/gpl-3.0.html
Uses Python libraries: mlxtend, pandas, networkx, matplotlib, openpyxl, flask, werkzeug
"""

import logging
import os
import uuid
import io
import base64

from flask import Flask, request, render_template, send_file, session
from flask_session import Session
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

import pandas as pd
from mlxtend.frequent_patterns import apriori, association_rules
from mlxtend.preprocessing import TransactionEncoder
import networkx as nx
import matplotlib.pyplot as plt

# ------------------------------------------------------
# Flask Configuration
# ------------------------------------------------------

app = Flask(__name__)

# Use a secure secret key for session (server-side)
app.secret_key = os.urandom(24)

# Server-side session config
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = os.path.join(app.root_path, 'flask_session')
app.config['SESSION_PERMANENT'] = False
Session(app)

# Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Folders and allowed extensions
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'xlsx'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(app.config['SESSION_FILE_DIR'], exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB

# ------------------------------------------------------
# Helper
# ------------------------------------------------------

def allowed_file(filename: str) -> bool:
    """Check if the uploaded file is .xlsx."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ------------------------------------------------------
# Routes
# ------------------------------------------------------

@app.route('/')
def index():
    """
    Main page: shows the upload form and thresholds.
    Retrieves thresholds & analysis results from session if available.
    """
    min_support = session.get('min_support', 0.01)
    min_confidence = session.get('min_confidence', 0.1)
    min_combination_count = session.get('min_combination_count', 1)

    rules_display = session.get('rules_display', None)
    download_link = session.get('download_link', None)
    download_graph_link = session.get('download_graph_link', None)
    graph_base64 = session.get('graph', None)
    no_rules = session.get('no_rules', False)

    return render_template(
        'upload.html',
        rules=rules_display,
        download_link=download_link,
        download_graph_link=download_graph_link,
        graph=graph_base64,
        min_support=min_support,
        min_confidence=min_confidence,
        min_combination_count=min_combination_count,
        no_rules=no_rules
    )

@app.route('/upload', methods=['POST'])
def upload_file():
    """
    Handle the file upload & thresholds, do Apriori analysis, 
    generate Excel & Graph, store data in session or on disk.
    """
    if 'file' not in request.files:
        logger.error("No file in request.")
        return render_template(
            'upload.html',
            error="No file in request.",
            min_support=0.01,
            min_confidence=0.1,
            min_combination_count=1,
            no_rules=False
        ), 400

    file = request.files['file']
    if file.filename == '':
        logger.error("No selected file.")
        return render_template(
            'upload.html',
            error="No selected file.",
            min_support=0.01,
            min_confidence=0.1,
            min_combination_count=1,
            no_rules=False
        ), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        logger.info(f"File saved: {file_path}")

        try:
            # Fetch thresholds
            min_support = float(request.form.get('min_support', 0.01))
            min_confidence = float(request.form.get('min_confidence', 0.1))
            min_combination_count = int(request.form.get('min_combination_count', 1))

            # Store in session
            session['min_support'] = min_support
            session['min_confidence'] = min_confidence
            session['min_combination_count'] = min_combination_count

            session['uploaded_file'] = file_path
            logger.info(
                f"Upload thresholds => min_support={min_support}, "
                f"min_confidence={min_confidence}, "
                f"min_combination_count={min_combination_count}"
            )

            # Read Excel
            df = pd.read_excel(file_path, sheet_name='Export', engine='openpyxl')
            logger.info("Excel read successfully.")

            grouped = df.groupby('Notification')['SPC'].apply(list).reset_index(name='Items')
            transactions = grouped['Items'].tolist()

            te = TransactionEncoder()
            te_ary = te.fit(transactions).transform(transactions)
            df_transformed = pd.DataFrame(te_ary, columns=te.columns_)

            total_transactions = len(transactions)
            logger.info(f"Total transactions: {total_transactions}")

            # Apriori
            frequent_itemsets = apriori(df_transformed, min_support=min_support, use_colnames=True)
            rules = association_rules(frequent_itemsets, metric="confidence", min_threshold=min_confidence)

            # Convert frozensets to strings
            rules['antecedents'] = rules['antecedents'].apply(lambda x: ', '.join(list(x)))
            rules['consequents'] = rules['consequents'].apply(lambda x: ', '.join(list(x)))

            # If zhangs_metric doesn't exist
            if 'zhangs_metric' not in rules.columns:
                rules['zhangs_metric'] = None

            # Combination count
            rules['count'] = (rules['support'] * total_transactions).round().astype(int)

            # Confidence to percent
            rules['confidence'] = (rules['confidence'] * 100).round(2)

            # Filter by minimal thresholds
            rules = rules[
                (rules['support'] >= min_support) &
                (rules['confidence'] >= min_confidence) &
                (rules['count'] >= min_combination_count)
            ]

            # Convert for display
            rules_display = rules[[
                'antecedents', 'consequents', 'count', 'support', 'confidence',
                'lift','leverage','conviction','zhangs_metric'
            ]].to_dict(orient='records')
            session['rules_display'] = rules_display

            if rules.empty:
                session['rules_display'] = []
                session['download_link'] = None
                session['download_graph_link'] = None
                session['graph'] = None
                session['no_rules'] = True
                logger.info("No rules found for given thresholds.")
                return render_template(
                    'upload.html',
                    rules=session['rules_display'],
                    download_link=None,
                    download_graph_link=None,
                    graph=None,
                    min_support=min_support,
                    min_confidence=min_confidence,
                    min_combination_count=min_combination_count,
                    no_rules=True
                )

            session['no_rules'] = False

            # Write rules to Excel in memory
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                rules[[
                    'antecedents', 'consequents', 'count', 'support', 'confidence',
                    'lift','leverage','conviction','zhangs_metric'
                ]].to_excel(writer, index=False)
            output.seek(0)

            # Format confidence
            wb = load_workbook(output)
            ws = wb.active
            confidence_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'confidence':
                    confidence_col = idx
                    break
            if confidence_col:
                for row in ws.iter_rows(min_row=2, min_col=confidence_col, max_col=confidence_col):
                    for cell in row:
                        cell.number_format = '0.00%'
            else:
                logger.warning("'confidence' column not found.")

            new_output = io.BytesIO()
            wb.save(new_output)
            new_output.seek(0)

            unique_id = uuid.uuid4().hex
            excel_filename = f'SPC_Rules_{unique_id}.xlsx'
            excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
            with open(excel_path, 'wb') as f:
                f.write(new_output.getbuffer())
            session['download_link'] = f'/download/{excel_filename}'

            # -------------------------------------------
            # Build Graph
            # -------------------------------------------
            G = nx.DiGraph()

            # Summiere Combination Count pro Knoten,
            # damit die Node-Size vom aufsummierten Count abhängt
            node_sizes_dict = {}

            for _, row in rules.iterrows():
                antecedents = row['antecedents'].split(', ')
                consequents = row['consequents'].split(', ')
                combo_count = row['count']

                # NODES
                for item in antecedents + consequents:
                    # Node hinzufügen
                    G.add_node(item)
                    # Summiere den count
                    if item not in node_sizes_dict:
                        node_sizes_dict[item] = 0
                    node_sizes_dict[item] += combo_count

                # EDGES
                confidence_val = row['confidence']
                lift_val = row['lift']
                for antecedent in antecedents:
                    for consequent in consequents:
                        if G.has_edge(antecedent, consequent):
                            G[antecedent][consequent]['weight'] += confidence_val
                            G[antecedent][consequent]['lift'] = (
                                G[antecedent][consequent]['lift'] + lift_val
                            ) / 2
                        else:
                            G.add_edge(
                                antecedent,
                                consequent,
                                weight=confidence_val,
                                lift=lift_val
                            )

            # Sortiere die Node-Sizes in der Reihenfolge von G.nodes()
            ordered_node_sizes = []
            for node in G.nodes():
                # Falls irgendwas schief geht: Default 50
                size_val = node_sizes_dict.get(node, 10)
                # z. B. * 10 als Skalierung
                ordered_node_sizes.append(size_val * 10)

            fig, ax = plt.subplots(figsize=(30, 30), dpi=500)
            pos = nx.spring_layout(G, k=2.9, iterations=200)

            # Node Size = ordered_node_sizes
            nx.draw_networkx_nodes(
                G,
                pos,
                node_size=ordered_node_sizes,
                node_color='lightblue',
                ax=ax
            )
            # Edges
            edges = G.edges(data=True)
            weights = [edge[2]['weight'] for edge in edges]
            nx.draw_networkx_edges(
                G, pos,
                arrowstyle='->',
                arrowsize=30,
                edge_color='gray',
                width=2,
                ax=ax
            )
            # Labels
            nx.draw_networkx_labels(G, pos, font_size=14, font_family='sans-serif', ax=ax)

            if weights:
                sm = plt.cm.ScalarMappable(cmap=plt.cm.Blues, norm=plt.Normalize(vmin=min(weights), vmax=max(weights)))
                sm.set_array([])
                cbar = fig.colorbar(sm, ax=ax)
                cbar.set_label('Confidence (%)')

            plt.title('Association Rules Network Graph', fontsize=20)
            plt.axis('off')

            # Save Graph to BytesIO
            img = io.BytesIO()
            plt.savefig(img, format='png', bbox_inches='tight')
            plt.close()

            # Speichere Graph-PNG
            graph_filename = f'Graph_{unique_id}.png'
            graph_path = os.path.join(OUTPUT_FOLDER, graph_filename)
            with open(graph_path, 'wb') as f:
                f.write(img.getbuffer())
            session['download_graph_link'] = f'/download/{graph_filename}'

            # Base64 for inline
            graph_base64 = base64.b64encode(img.getvalue()).decode()
            session['graph'] = graph_base64

            return render_template(
                'upload.html',
                rules=rules_display,
                download_link=session['download_link'],
                download_graph_link=session['download_graph_link'],
                graph=graph_base64,
                min_support=min_support,
                min_confidence=min_confidence,
                min_combination_count=min_combination_count,
                no_rules=False
            )

        except Exception as e:
            logger.exception("Error while processing the file.")
            return render_template(
                'upload.html',
                error=f"An error occurred: {str(e)}",
                min_support=0.01,
                min_confidence=0.1,
                min_combination_count=1,
                no_rules=False
            ), 500
    else:
        logger.error("Invalid file format (not .xlsx).")
        return render_template(
            'upload.html',
            error="Invalid file format. Please upload a .xlsx.",
            min_support=0.01,
            min_confidence=0.1,
            min_combination_count=1,
            no_rules=False
        ), 400

@app.route('/refresh', methods=['POST'])
def refresh():
    """
    Refresh the analysis with updated minimal thresholds (no new upload).
    """
    file_path = session.get('uploaded_file', None)
    if not file_path or not os.path.exists(file_path):
        logger.error("No uploaded file found. Please upload a file first.")
        return render_template(
            'upload.html',
            error="No uploaded file found. Please upload a file first.",
            min_support=0.01,
            min_confidence=0.1,
            min_combination_count=1,
            no_rules=False
        ), 400

    try:
        min_support = float(request.form.get('min_support', 0.01))
        min_confidence = float(request.form.get('min_confidence', 0.1))
        min_combination_count = int(request.form.get('min_combination_count', 1))

        session['min_support'] = min_support
        session['min_confidence'] = min_confidence
        session['min_combination_count'] = min_combination_count

        logger.info(
            f"Refresh -> min_support={min_support}, "
            f"min_confidence={min_confidence}, "
            f"min_combination_count={min_combination_count}"
        )

        df = pd.read_excel(file_path, sheet_name='Export', engine='openpyxl')
        grouped = df.groupby('Notification')['SPC'].apply(list).reset_index(name='Items')
        transactions = grouped['Items'].tolist()

        te = TransactionEncoder()
        te_ary = te.fit(transactions).transform(transactions)
        df_transformed = pd.DataFrame(te_ary, columns=te.columns_)

        total_transactions = len(transactions)
        frequent_itemsets = apriori(df_transformed, min_support=min_support, use_colnames=True)
        rules = association_rules(frequent_itemsets, metric="confidence", min_threshold=min_confidence)

        rules['antecedents'] = rules['antecedents'].apply(lambda x: ', '.join(list(x)))
        rules['consequents'] = rules['consequents'].apply(lambda x: ', '.join(list(x)))
        if 'zhangs_metric' not in rules.columns:
            rules['zhangs_metric'] = None

        rules['count'] = (rules['support'] * total_transactions).round().astype(int)
        rules['confidence'] = (rules['confidence'] * 100).round(2)

        rules = rules[
            (rules['support'] >= min_support) &
            (rules['confidence'] >= min_confidence) &
            (rules['count'] >= min_combination_count)
        ]

        rules_display = rules[[
            'antecedents', 'consequents', 'count', 'support', 'confidence',
            'lift','leverage','conviction','zhangs_metric'
        ]].to_dict(orient='records')
        session['rules_display'] = rules_display

        if rules.empty:
            session['rules_display'] = []
            session['download_link'] = None
            session['download_graph_link'] = None
            session['graph'] = None
            session['no_rules'] = True
            logger.info("No rules found for the updated thresholds.")
            return render_template(
                'upload.html',
                rules=session['rules_display'],
                download_link=None,
                download_graph_link=None,
                graph=None,
                min_support=min_support,
                min_confidence=min_confidence,
                min_combination_count=min_combination_count,
                no_rules=True
            )

        session['no_rules'] = False

        # We do not regenerate Excel/Graph here (unless you want to do so).
        return render_template(
            'upload.html',
            rules=rules_display,
            download_link=session.get('download_link', None),
            download_graph_link=session.get('download_graph_link', None),
            graph=session.get('graph', None),
            min_support=min_support,
            min_confidence=min_confidence,
            min_combination_count=min_combination_count,
            no_rules=False
        )

    except Exception as e:
        logger.exception("Error while refreshing analysis.")
        return render_template(
            'upload.html',
            error=f"An error occurred: {str(e)}",
            min_support=0.01,
            min_confidence=0.1,
            min_combination_count=1,
            no_rules=False
        ), 500

@app.route('/download/<filename>')
def download_file_route(filename):
    """
    Download any file (Excel or PNG) from the outputs folder.
    """
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    if not os.path.exists(filepath):
        logger.error(f"Requested file not found: {filepath}")
        return "File not found.", 404

    return send_file(filepath, as_attachment=True)

if __name__ == '__main__':
    # Start Flask in debug mode
    app.run(host='0.0.0.0', port=5000, debug=True)
